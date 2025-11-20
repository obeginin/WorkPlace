import pandas as pd
import json
import os
from typing import List, Dict, Any
from pathlib import Path
import logging
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from typing import Union, List
import csv
import logging
from ClassLogger import LoggerConfig
from ClassFiles import FileManager


class DataConverter:
    """Класс для конвертации данных между различными форматами"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def _log_info(self, message: str) -> None:
        self.logger.info(f"FileManager - {message}")

    def _log_error(self, message: str) -> None:
        self.logger.error(f"FileManager - {message}")

    def txt_to_csv(self, input_file: str, chunk_size: int = 100000):
        """Конвертирует txt файл (каждая строка - одно имя) в CSV"""
        input_path = Path(input_file)
        output_file = input_path.parent / f"{input_path.stem}.csv"
        processed = 0
        encoding: str = 'utf-8'
        try:
            with open(output_file, 'a', newline='', encoding=encoding) as outfile:
                writer = csv.writer(outfile)

                # Записываем заголовки с добавлением столбца №
                writer.writerow(['№', 'filename'])

                # Читаем и обрабатываем частями
                buffer = []
                with open(input_path, 'r', encoding=encoding) as infile:
                    for line_num, line in enumerate(infile, 1):  # начинаем нумерацию с 1
                        filename = line.strip()
                        if filename:  # пропускаем пустые строки
                            # Добавляем номер строки и имя файла
                            buffer.append([line_num, filename])
                            processed += 1

                            # Сбрасываем буфер при достижении размера
                            if len(buffer) >= chunk_size:
                                writer.writerows(buffer)
                                buffer.clear()
                                self._log_info(f"Обработано: {processed:,} строк")

                # Записываем остаток
                if buffer:
                    writer.writerows(buffer)

            self._log_info(f"✅ Готово! Обработано строк: {processed:,}")
            self._log_info(f"📁 Файл: {output_file}")

        except Exception as e:
            self._log_error(f"❌ Ошибка: {e}")

    def txt_to_csv_chunked(self, input_file: str, chunk_size: int = 100_000):
        """Конвертирует txt файл в csv частями, используя генератор."""
        input_path = Path(input_file)
        output_file = input_path.with_suffix(".csv")
        encoding = "utf-8"
        total_processed = 0
        line_offset = 0
        try:
            # Удаляем старый CSV, если есть
            if output_file.exists():
                output_file.unlink()
            # Пишем заголовок один раз
            with open(output_file, "w", newline="", encoding=encoding) as f:
                writer = csv.writer(f)
                writer.writerow(["№", "filename"])

            # Обрабатываем файл порциями (исходный файл)
            big_file_txt = FileManager()
            for chunk in big_file_txt.read_large_file_chunked(input_path, chunk_size=chunk_size, encoding=encoding):
                rows = []
                for i, line in enumerate(chunk, 1):
                    if line.strip():
                        rows.append([line_offset + i, line.strip()])
                # Добавляем строки в CSV
                with open(output_file, "a", newline="", encoding=encoding) as f:
                    writer = csv.writer(f)
                    writer.writerows(rows)
                total_processed += len(rows)
                line_offset += len(chunk)
                self._log_info(f"Обработано: {total_processed:,} строк")
            self._log_info(f" Готово! Всего обработано: {total_processed:,} строк")
            self._log_info(f" Файл: {output_file}")
        except Exception as e:
            self._log_error(f" Ошибка при конвертации: {e}")



    def txt_to_csv_large(self, input_file: str, chunk_size: int = 100000):
        """Конвертирует txt файл (каждая строка - одно имя) в CSV"""
        input_path = Path(input_file)
        output_file = input_path.parent / f"{input_path.stem}.csv"
        processed = 0
        encoding: str = 'utf-8'

        try:
            big_file = FileManager()

            with open(output_file, 'a', newline='', encoding=encoding) as outfile:
                writer = csv.writer(outfile)

                # Записываем заголовки с добавлением столбца №
                writer.writerow(['№', 'filename'])

                # Читаем и обрабатываем частями
                buffer = []
                for chunk in big_file.read_large_file_chunked():
                    for line_num, line in enumerate(chunk, 1):

                        filename = line.strip()
                        if filename:
                            buffer.append([line_num, filename])
                            processed += 1

                            # Сбрасываем буфер при достижении размера
                            if len(buffer) >= chunk_size:
                                writer.writerows(buffer)
                                buffer.clear()
                                self._log_info(f"Обработано: {processed:,} строк")

                # Записываем остаток
                if buffer:
                    writer.writerows(buffer)

            self._log_info(f"✅ Готово! Обработано строк: {processed:,}")
            self._log_info(f"📁 Файл: {output_file}")

        except Exception as e:
            self._log_error(f"❌ Ошибка: {e}")

    def json_to_excel(self, input_file: str, output_file: str) -> pd.DataFrame:
        """Конвертирует JSON файл (каждая строка - отдельный JSON) в Excel"""
        data = []
        with open(input_file, "r", encoding="utf-8") as file:
            for line_num, line in enumerate(file, 1):
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                    data.append(entry)
                except json.JSONDecodeError as e:
                    self._log_info(f"⚠️ Ошибка в строке {line_num}: {e}")
                    self._log_info(f"   Проблемная строка: {line[:100]}...")
                    continue

        self._log_info.info(f"📊 Обработано записей: {len(data)}")
        df = pd.DataFrame(data)

        # Создаем директорию если нет
        os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)

        df.to_excel(output_file, index=False, engine='openpyxl')
        self._log_info(f"✅ Готово! Данные сохранены в {output_file}")
        return df

    def json_to_txt(self, input_file: str, output_file: str, delimiter: str = " | ") -> None:
        """
        Конвертирует JSON файл в текстовый формат
        """
        data = []
        with open(input_file, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                    data.append(entry)
                except json.JSONDecodeError:
                    continue

        if not data:
            self._log_info("⚠️ Нет данных для конвертации")
            return

        df = pd.DataFrame(data)

        with open(output_file, "w", encoding="utf-8") as f:
            # Записываем заголовки
            headers = delimiter.join(df.columns)
            f.write(headers + "\n")

            # Записываем данные
            for _, row in df.iterrows():
                line = delimiter.join(str(row[col]) for col in df.columns)
                f.write(line + "\n")

        self._log_info(f"✅ Готово! Данные сохранены в {output_file}")


    def json_to_python(self, input_file: str) -> List[Dict[str, Any]]:
        """Конвертирует JSON файл (каждая строка - отдельный JSON) в Excel"""
        data = []
        with open(input_file, "r", encoding="utf-8") as file:
            for line_num, line in enumerate(file, 1):
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                    data.append(entry)
                except json.JSONDecodeError as e:
                    self._log_info(f"⚠️ Ошибка в строке {line_num}: {e}")
                    self._log_info(f"   Проблемная строка: {line[:100]}...")
                    continue

        self._log_info(f"📊 Загружено {len(data)} записей из файла {input_file}")
        return data

    def python_to_excel(self, data: Dict[Any, Dict], output_file: str='template.xlsx', key_name: str = "key"):
        '''Конвертирует словарь в Excel таблицу
        Args:
            data_dict: Словарь {ключ: {данные}}
            output_file: Путь для сохранения Excel файла
            key_name: Название колонки для ключей (по умолчанию "key")'''

        # Собираем все данные в список словарей
        rows = []
        for key, inner_dict in data.items():
            row = {key_name: key}
            row.update(inner_dict)
            rows.append(row)

        df = pd.DataFrame.from_dict(rows)

        # Обрабатываем списки и множества - преобразуем в строки
        for column in df.columns:
            if df[column].apply(lambda x: isinstance(x, (list, set, dict))).any():
                df[column] = df[column].apply(
                    lambda x: ', '.join(map(str, x)) if isinstance(x, (list, set)) else str(x))

        output_file = fr'C:\Users\beginin-ov\Projects\Local\files\{output_file}'
        df.to_excel(output_file, index=False, engine='openpyxl')
        self._log_info(f"✅ Данные сохранены в {output_file}")
        self._log_info(f"📊 Структура таблицы: {len(df)} строк, {len(df.columns)} колонок: {list(df.columns)}")

    def read_txt_file(self, file_path: str, encoding: str = "utf-8") -> list[str]:
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                lines = [line.strip() for line in file if line.strip()]
            self._log_info(f"✅ Прочитано {len(lines)} строк из {file_path}")
            return lines
        except FileNotFoundError:
            self._log_info(f"❌ Файл не найден: {file_path}")
            return []
        except Exception as e:
            self._log_info(f"❌ Ошибка чтения файла {file_path}: {e}")
            return []
        #return df


    def python_to_excel_with_id(self, data: List[Dict], output_file: str = 'template.xlsx', add_id: bool = True):
        '''Конвертирует список словарей в Excel с автоматическим ID'''
        if not data:
            self._log_info("❌ Передан пустой список данных")
            return

        df = pd.DataFrame(data)

        # Добавляем колонку с ID если нужно
        if add_id and 'id' not in df.columns:
            df.insert(0, 'id', range(1, len(df) + 1))

        # Обработка сложных типов
        for column in df.columns:
            if df[column].apply(lambda x: isinstance(x, (list, set, dict))).any():
                df[column] = df[column].apply(
                    lambda x: ', '.join(map(str, x)) if isinstance(x, (list, set)) else str(x)
                )

        output_path = fr'C:\Users\beginin-ov\Projects\Local\files\{output_file}'
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        df.to_excel(output_path, index=False, engine='openpyxl')

        self._log_info(f"✅ Данные сохранены в {output_path}")
        self._log_info(f"📊 Структура: {len(df)} строк, {len(df.columns)} колонок")

        return df

    def txt_to_excel_optimized(self, input_file: str, output_file: str, chunk_size: int = 20000) -> None:
        """
        Конвертирует обычный текстовый файл в Excel
        Каждая строка текста становится отдельной записью
        """
        from tqdm import tqdm

        # Сначала считаем общее количество строк
        self._log_info("📊 Подсчет общего количества строк...")
        total_lines = 0
        with open(input_file, "r", encoding="utf-8") as file:
            for _ in file:
                total_lines += 1

        self._log_info(f"📁 Всего строк в файле: {total_lines}")

        # Создаем Excel writer
        output_path = fr'C:\Users\beginin-ov\Projects\Local\files\{output_file}'
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        all_data = []
        processed_lines = 0

        self._log_info("📝 Обработка текстового файла...")

        with open(input_file, "r", encoding="utf-8") as file:
            with tqdm(total=total_lines, desc="Обработка") as pbar:
                for line_num, line in enumerate(file, 1):
                    line = line.strip()

                    if line:  # Только непустые строки
                        # Создаем запись для каждой строки
                        record = {
                            "line_number": line_num,
                            "text": line,
                            "length": len(line)
                        }
                        all_data.append(record)
                        processed_lines += 1

                    # Записываем чанками для экономии памяти
                    if len(all_data) >= chunk_size:
                        df = pd.DataFrame(all_data)
                        if line_num == chunk_size:  # Первый чанк
                            df.to_excel(output_path, index=False, engine='openpyxl')
                        else:  # Последующие чанки
                            with pd.ExcelWriter(output_path, mode='a', engine='openpyxl',
                                                if_sheet_exists='overlay') as writer:
                                from openpyxl import load_workbook
                                wb = load_workbook(output_path)
                                ws = wb.active
                                last_row = ws.max_row
                                df.to_excel(writer, index=False, header=False, startrow=last_row)

                        all_data = []  # Очищаем для следующего чанка
                        self._log_info(f"📦 Записан чанк с {chunk_size} строками")

                    pbar.update(1)

        # Записываем остаток данных
        if all_data:
            df = pd.DataFrame(all_data)
            if processed_lines <= chunk_size:  # Если данные уместились в один чанк
                df.to_excel(output_path, index=False, engine='openpyxl')
            else:  # Добавляем к существующему файлу
                with pd.ExcelWriter(output_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    from openpyxl import load_workbook
                    wb = load_workbook(output_path)
                    ws = wb.active
                    last_row = ws.max_row
                    df.to_excel(writer, index=False, header=False, startrow=last_row)

        self._log_info(f"✅ Файл сохранен: {output_path}")
        self._log_info(f"📊 Обработано строк: {processed_lines}")

    def ensure_dir_exists(self, path: Path) -> None:
        """Создает директорию, если её нет"""
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            self._log_info(f"Не удалось создать директорию для {path}: {e}")
            raise

    def txt_to_xlsx_stream(
            self,
            input_path: Union[str, Path],
            output_path: Union[str, Path],
            buffer_size: int = 10_000,
            separator: str = None,
            show_progress: bool = True
    ) -> bool:
        """
        Преобразует большой текстовый файл в XLSX, с прогресс-баром.
        """
        input_path = Path(input_path)
        output_path = Path(output_path)
        self.ensure_dir_exists(output_path)

        try:
            total_size = os.path.getsize(input_path)
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Data")

            with open(input_path, "r", encoding="utf-8") as f:
                buffer: List[List[str]] = []
                count = 0
                last_percent = -1

                # Читаем файл порциями для отслеживания прогресса
                while True:
                    line = f.readline()
                    if not line:  # конец файла
                        break

                    line = line.strip()
                    if not line:
                        continue

                    row = line.split(separator) if separator else [line]
                    buffer.append(row)
                    count += 1

                    if len(buffer) >= buffer_size:
                        for r in buffer:
                            ws.append(r)
                        buffer.clear()

                    if show_progress:
                        pos = f.tell()
                        percent = int((pos / total_size) * 100)
                        if percent != last_percent:
                            sys.stdout.write(f"\rПрогресс: {percent}% ({count:,} строк)")
                            sys.stdout.flush()
                            last_percent = percent

                # финальный сброс
                if buffer:
                    for r in buffer:
                        ws.append(r)

            wb.save(output_path)
            if show_progress:
                self._log_info(f"\r Готово! Файл {input_path.name} успешно преобразован в {output_path.name}. Всего строк: {count:,}")
            return True

        except Exception as e:
            self._log_error(f"Ошибка при конвертации {input_path} в XLSX: {e}")
            return False


            
#converter = DataConverter()
#converter.filenames_txt_to_csv(input_file=r"C:\Users\beginin-ov\Projects\Local\files\CP_succes.txt", output_file="files_list.csv")
#converter.filenames_txt_to_csv(input_file=r"C:\Users\beginin-ov\Projects\Local\files\DLIVR.txt", output_file="DLIVR.csv")
#converter.json_to_excel(input_file="2025-10-23.json",output_file="AppSimChecher_2025-10-23.xlsx")
#converter.json_to_excel(input_file="analys_log_long.json",output_file="analys_long.xlsx")
#list_from_txt = converter.read_txt_file(file_path=r"C:\Users\beginin-ov\Projects\Local\files\results\CP_succes.txt")

#'''converter.txt_to_excel_optimized(
#    input_file=r"C:\Users\beginin-ov\Projects\Local\files\results\DLIVR_succes.txt",
#    output_file="DLIVR_succes.xlsx",
#    chunk_size=50000
#)'''
#converter.python_to_excel_with_id(data=list_from_txt, output_file="CP_succes.xlsx")



# converter.txt_to_xlsx_stream(
#     input_path=r"C:\Users\beginin-ov\Projects\Local\files\results_2\DLAPI_succes — копия (4).txt",
#     output_path="DLAPI_succes_new4.xlsx",
#     buffer_size=50000
# )
# converter.txt_to_xlsx_stream(
#     input_path=r"C:\Users\beginin-ov\Projects\Local\files\results_2\DLAPI_succes — копия (5).txt",
#     output_path="DLAPI_succes_new5.xlsx",
#     buffer_size=50000
# )
#
# converter.txt_to_xlsx_stream(
#     input_path=r"C:\Users\beginin-ov\Projects\Local\files\results_2\DLAPI_succes — копия (6).txt",
#     output_path="DLAPI_succes_new6.xlsx",
#     buffer_size=50000
# )
# converter.txt_to_xlsx_stream(
#     input_path=r"C:\Users\beginin-ov\Projects\Local\files\results_2\DLAPI_succes — копия (7).txt",
#     output_path="DLAPI_succes_new7.xlsx",
#     buffer_size=50000
# )
#
# converter.txt_to_xlsx_stream(
#     input_path=r"C:\Users\beginin-ov\Projects\Local\files\results_2\DLAPI_succes — копия (8).txt",
#     output_path="DLAPI_succes_new8.xlsx",
#     buffer_size=50000
# )

#fm = FileManager(logger= logger)
#for chunk in fm.read_large_file_chunked(r"C:\Users\beginin-ov\Projects\Local\files\CP_succes.txt"):

    #(f"прог")